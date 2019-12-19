import openpyxl
import os



dict_list = []
# Записываем значения словаря в СПИСОК dict_list
# Ниже блок WITH
with open('dictionary.txt') as f_dictionary:
     for word_line in f_dictionary: dict_list.append(word_line.split("\n")[0])


all_files = os.listdir() # Читаем все файлы в директории в список
xls_files = [] # Объявляем список куда сложим все файлы xlsx

# Из всего списка файлов формируем список файло xls
# НАЧАЛО
for i in range(0,len(all_files)):
    if all_files[i].endswith('.xlsx') and all_files[i].find("_1") == -1:
        xls_files.append(all_files[i])
# КОНЕЦ

for tmp_file in xls_files:
    wb = openpyxl.load_workbook(filename = tmp_file)
    for k in wb.sheetnames: # Пербор вкладок книги Excel

        ws=wb[k] # Читаем в ws вкладку

        end_row = ws.max_row + 1 # Для полной обработки файла

        # end_row = ws.max_row + 1 # Для полной обработки файла
        # end_row = ws.max_row + 1 # Для полной обработки файла
        end_column = ws.max_column + 1

        for i in range(1,end_row):

            # Количество столбцов в которых встречаются значения из словаря. Изначально 0. Переменная нужна для сдвигая записи совпадающих значений найденных в разных столбцах
            # ниже одна строка
            occurrence_number = 0

            for j in range(1, end_column):
                cell_value = str(ws.cell(row=i, column=j).value) # Читаем в переменную Значение текущей ячейки
                cell_name =  str(ws.cell(row=i, column=j))       # Читаем в переменную Имя текущей ячейки


                # Количество встречающихся слов из словаря. Первое совпадение пишется в ячейку,
                # последующие добавляются. Чтобы реже писать в ячейку слово "Нашел !!!".
                # Чтобы реже
                # ниже одна строка
                dict_number = 0

                for dict_line_tmp in dict_list:
                    # Блок FOR переборает все значения словаря в текущей ячейке

                    if cell_value.lower().find(dict_line_tmp) != -1:
                        if dict_number == 0:
                            # БЛОК IF Выполняется если в ячейке найдено первое совпадение из словаря

                            ws.cell(row=i, column=end_column).value = "Нашел !!!"

                            # Запись в ячейку имени ячейки в которой встречается значение из словаря
                            # ниже одна строка
                            ws.cell(row=i, column=end_column + occurrence_number + 1).value = cell_name



                            # Запись в ячейку найденного значение из словаря
                            # ниже одна строка
                            ws.cell(row=i, column=end_column + occurrence_number + 2).value = dict_line_tmp

                            # Выставляем признак того что первое слово словаря найдено.
                            # Дальше по итерациям нужно не присваивать а добавлять к ячейке.
                            # Плюс далее используется как признак что запись о нахождении совпадений нужно делать
                            # в столбцы на 2 значения правее
                            # ниже одна строка
                            dict_number = 1
                        else:
                            # БЛОК ELSE Выполняется если найдены последующие совпадения из словаря (более одного)

                            # Запись в ячейку имени ячейки в которой встречается значение из словаря
                            # ниже одна строка
                            ws.cell(row=i, column=end_column + occurrence_number + 1).value = cell_name


                            # Запись в ячейку найденного значение из словаря
                            # ниже одна строка
                            ws.cell(row=i, column=end_column + occurrence_number + 2).value = str(ws.cell(row=i, column=end_column + occurrence_number + 2).value) + ", \n" +dict_line_tmp

                # Сдвигаем запись найденных слов словаря на 2 в право. Происходит только если в текущей ячейке было найдено хотя бы одно значение из словаря.
                # одна строка
                if dict_number == 1 : occurrence_number += 2


                # Печать текущих строки и строчки которые обрабатываются
                # одна строка
                print (tmp_file," ", wb[k]," ", i," ",j)


    # Готовим имя выходного файла. Для этого вырезаем имя исходного (без расшрения)
    # одна строка
    just_first_part_of_file_name = tmp_file.split(".xlsx")[0]

    # Формируем имя выходного файла. Имя выходного файла = имя исходного + "_1"
    # одна строка
    name_with_1_and_xls = just_first_part_of_file_name + "_1.xlsx"

    # Проверка и удаление в папке файла с именем "_1", под которым мы планируем сохранять
    # одна строка
    if os.path.exists(name_with_1_and_xls): os.remove(name_with_1_and_xls)

    # Сохраняем файл результат
    # На больших файлах может возникать ошибка поэтому отлавливаем исключением сохранение файла
    # Блок TRY
    try:
        wb.save(name_with_1_and_xls)  # Для сохранения нужна только эта операция
    except Exception:
        print('Ошибка при сохранении файла ', name_with_1_and_xls)
    else:
        print('Файл ', name_with_1_and_xls, " сохранен")



