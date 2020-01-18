# Разбивает исходные xlsx файлы на файлы по 5000 строчек (current_first_line)

import openpyxl
from openpyxl import Workbook
import os


all_files = os.listdir() # Читаем все файлы в директории в список
xls_files = [] # Объявляем список куда сложим все файлы xlsx

# Из всего списка файлов формируем список файло xls
# НАЧАЛО

for i in range(0,len(all_files)):
    if all_files[i].endswith('.xlsx') and all_files[i].find("result") == -1:
        xls_files.append(all_files[i])
# КОНЕЦ

print (xls_files)

for tmp_file in xls_files:
    wb = openpyxl.load_workbook(filename=tmp_file)
    first_sheet_name = wb.sheetnames[0]
    ws=wb[first_sheet_name]
    end_row = ws.max_row + 1
    end_column = ws.max_column + 1

    current_first_line = 1
    current_name = 1

    for i in range(1,end_row):

        print("Обработка строки " + str(i))

        # Если i соответствует следующей новой строке то нужно открыть новый файл
        if i == current_first_line:
            wb_to_save = Workbook()
            ws_to_save = wb_to_save[wb_to_save.sheetnames[0]]
            current_first_line += 5000 # количество строк в результируещем файле
            k = 1
            m = 1

        for j in range(1, end_column):

            ws_to_save.cell(row=k, column=m).value =ws.cell(row=i, column=j).value
            m += 1

        m = 1
        k += 1




        if i == current_first_line - 1:
            just_first_part_of_file_name = tmp_file.split(".xlsx")[0]
            name_with_index_and_xls = just_first_part_of_file_name + '_result_' + str(current_name) + '.xlsx'
            if os.path.exists(name_with_index_and_xls): os.remove(name_with_index_and_xls)

            wb_to_save.save(name_with_index_and_xls)
            current_name += 1

    if i != current_first_line - 1:
        just_first_part_of_file_name = tmp_file.split(".xlsx")[0]
        name_with_index_and_xls = just_first_part_of_file_name + '_result_' + str(current_name) + '.xlsx'
        if os.path.exists(name_with_index_and_xls): os.remove(name_with_index_and_xls)

        wb_to_save.save(name_with_index_and_xls)
        current_name += 1




